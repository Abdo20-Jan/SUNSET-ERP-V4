#!/usr/bin/env python3
"""Genera el dataset TS del plan de cuentas a partir del Excel maestro.

Lee `PLANO DE CONTAS FINAL.xlsx` (aba `plano de contas`, 19 columnas × 631
cuentas) y emite `src/lib/services/plan-de-cuentas.data.ts`: una entrada
`CuentaPlan` por cuenta, fiel al Excel. `nivel`/`padreCodigo`/`categoria`/
`moneda` NO se emiten — se derivan en `planEntryToSeedRecord`.

Uso:
  python3 scripts/extract-plan-de-cuentas.py <ruta.xlsx> [--check]
    --check  : sólo valida invariantes e imprime el reporte; no escribe.
"""

from __future__ import annotations

import sys
from collections import Counter, defaultdict

import openpyxl

CLASIFICACION = {
    "Activo": "ACTIVO",
    "Pasivo": "PASIVO",
    "Corriente": "CORRIENTE",
    "No corriente": "NO_CORRIENTE",
    "Patrimonio neto": "PATRIMONIO_NETO",
    "Resultado": "RESULTADO",
}
TIPO = {"Sintética": "SINTETICA", "Analítica": "ANALITICA"}
NATURALEZA = {
    "DEUDORA": "DEUDOR",
    "ACREEDORA": "ACREEDOR",
    "MIXTA": "MIXTA",
    "SISTEMA / VARIABLE": "SISTEMA_VARIABLE",
}
IMPUTACION = {
    "Imputable": "IMPUTABLE",
    "No imputable": "NO_IMPUTABLE",
    "Solo sistema / cierre": "SOLO_SISTEMA",
}
SI_NO = {"Sí": True, "No": False}
# Categoría legada (ecuación patrimonial) derivada de la clase contable 1-9.
CATEGORIA_POR_CLASE = {
    1: "ACTIVO", 2: "PASIVO", 3: "PATRIMONIO", 4: "INGRESO",
    5: "EGRESO", 6: "EGRESO", 7: "EGRESO", 8: "EGRESO", 9: "EGRESO",
}


def limpiar_nombre(s: str) -> str:
    return str(s).replace(" ", "").strip()


def leer(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["plano de contas"]
    H = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=H.index(h) + 1).value for h in H}
        codigo = str(row["Código"]).strip()
        out.append({
            "orden": int(row["Orden"]),
            "codigo": codigo,
            "nombre": limpiar_nombre(row["Cuenta"]),
            "clase": int(str(row["Clase"]).split(" ")[0]),
            "clasificacion": CLASIFICACION[row["Clasificación"]],
            "tipo": TIPO[row["Tipo"]],
            "naturaleza": NATURALEZA[row["Naturaleza"]],
            "imputacion": IMPUTACION[row["Imputación"]],
            "regularizadora": SI_NO[row["Regularizadora"]],
            "bimonetaria": SI_NO[row["Bimonetaria"]],
            "monedaExtranjera": SI_NO[row["Moneda extranjera"]],
            "enEspecie": SI_NO[row["En especie"]],
            "inventariable": SI_NO[row["Inventariable"]],
            "sistema": SI_NO[row["Sistema"]],
            "dinamica": SI_NO[row["Dinámica"]],
            "_nivel_xls": int(row["Nivel"]),
            "_padre_xls": (str(row["Código padre"]).strip()
                           if row["Código padre"] is not None else None),
        })
    return out


def check(rows: list[dict]) -> int:
    cods = [r["codigo"] for r in rows]
    cset = set(cods)
    errs = []
    # 1) código único
    dup = [c for c, n in Counter(cods).items() if n > 1]
    if dup:
        errs.append(f"códigos duplicados: {dup}")
    # 2) orden único
    od = [o for o, n in Counter(r["orden"] for r in rows).items() if n > 1]
    if od:
        errs.append(f"orden duplicado: {od}")
    # 3) nivel derivado == Excel
    bad_n = [r["codigo"] for r in rows if r["codigo"].count(".") + 1 != r["_nivel_xls"]]
    if bad_n:
        errs.append(f"nivel derivado != Excel: {bad_n[:5]}")
    # 4) padre derivado == Excel y existe
    for r in rows:
        segs = r["codigo"].split(".")
        der = ".".join(segs[:-1]) if len(segs) > 1 else None
        if der != r["_padre_xls"]:
            errs.append(f"padre derivado != Excel en {r['codigo']}: {der} vs {r['_padre_xls']}")
        if der is not None and der not in cset:
            errs.append(f"huérfana {r['codigo']} → padre inexistente {der}")
    # 5) clase == 1er dígito
    bad_cl = [r["codigo"] for r in rows if r["clase"] != int(r["codigo"].split(".")[0])]
    if bad_cl:
        errs.append(f"clase != 1er dígito: {bad_cl[:5]}")
    # 6) categoría derivable para toda clase
    bad_cat = [r["codigo"] for r in rows if r["clase"] not in CATEGORIA_POR_CLASE]
    if bad_cat:
        errs.append(f"clase sin categoría: {bad_cat[:5]}")

    sint = {r["codigo"] for r in rows if r["tipo"] == "SINTETICA"}
    # 7) padre debe ser SINTETICA
    for r in rows:
        segs = r["codigo"].split(".")
        if len(segs) > 1 and ".".join(segs[:-1]) not in sint:
            errs.append(f"padre no sintético: {r['codigo']}")
    # 8) tipo × imputación
    for r in rows:
        if r["tipo"] == "SINTETICA" and r["imputacion"] == "IMPUTABLE":
            errs.append(f"sintética imputable: {r['codigo']}")
        if r["tipo"] == "ANALITICA" and r["imputacion"] == "NO_IMPUTABLE":
            errs.append(f"analítica no-imputable: {r['codigo']}")

    print("=== INVARIANTES ===")
    print("Total cuentas:", len(rows))
    print("Inventariable=true por clase:", dict(Counter(
        r["clase"] for r in rows if r["inventariable"])))
    print("Inventariable fuera de ACTIVO (clase!=1):",
          [r["codigo"] for r in rows if r["inventariable"] and r["clase"] != 1] or "NINGUNA")
    mt = defaultdict(Counter)
    for r in rows:
        mt[r["tipo"]][r["imputacion"]] += 1
    print("Tipo × Imputación:", {k: dict(v) for k, v in mt.items()})
    print("Naturaleza dist:", dict(Counter(r["naturaleza"] for r in rows)))
    print("Clasificación dist:", dict(Counter(r["clasificacion"] for r in rows)))
    print("Categoría derivada dist:", dict(Counter(
        CATEGORIA_POR_CLASE[r["clase"]] for r in rows)))
    # regularizadora con naturaleza vs categoría derivada
    reg_bad = []
    for r in rows:
        if r["regularizadora"]:
            cat = CATEGORIA_POR_CLASE[r["clase"]]
            default = "DEUDOR" if cat in ("ACTIVO", "EGRESO") else "ACREEDOR"
            if r["naturaleza"] in ("DEUDOR", "ACREEDOR") and r["naturaleza"] == default:
                reg_bad.append(r["codigo"])
    print("Regularizadoras con naturaleza = default (no invertida):", reg_bad or "NINGUNA")

    print("\n=== RESULTADO:", "OK ✓" if not errs else f"{len(errs)} ERRORES ✗", "===")
    for e in errs[:40]:
        print("  ✗", e)
    return 0 if not errs else 1


def emit_ts(rows: list[dict]) -> str:
    def b(v):
        return "true" if v else "false"

    lines = []
    for r in rows:
        nombre = r["nombre"].replace("\\", "\\\\").replace('"', '\\"')
        lines.append(
            "  { "
            f'orden: {r["orden"]}, codigo: "{r["codigo"]}", nombre: "{nombre}", '
            f'clase: {r["clase"]}, clasificacion: "{r["clasificacion"]}", '
            f'tipo: "{r["tipo"]}", naturaleza: "{r["naturaleza"]}", '
            f'imputacion: "{r["imputacion"]}", '
            f'regularizadora: {b(r["regularizadora"])}, bimonetaria: {b(r["bimonetaria"])}, '
            f'monedaExtranjera: {b(r["monedaExtranjera"])}, enEspecie: {b(r["enEspecie"])}, '
            f'inventariable: {b(r["inventariable"])}, sistema: {b(r["sistema"])}, '
            f'dinamica: {b(r["dinamica"])} }},'
        )
    body = "\n".join(lines)
    return (
        "/**\n"
        " * Dataset del plan de cuentas — GENERADO por `scripts/extract-plan-de-cuentas.py`\n"
        " * desde `PLANO DE CONTAS FINAL.xlsx` (631 cuentas × 19 columnas). NO editar a mano:\n"
        " * regenerar desde el Excel maestro. `nivel`/`padreCodigo`/`categoria`/`moneda` se\n"
        " * derivan en `planEntryToSeedRecord` (no se persisten acá).\n"
        " */\n"
        'import type { CuentaPlan } from "./plan-de-cuentas";\n\n'
        "// biome-ignore format: dataset generado — una cuenta por línea\n"
        "export const PLAN_CUENTAS_DATA: readonly CuentaPlan[] = [\n"
        f"{body}\n"
        "];\n"
    )


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("uso: extract-plan-de-cuentas.py <ruta.xlsx> [--check]")
        return 2
    rows = leer(args[0])
    rc = check(rows)
    if "--check" in sys.argv:
        return rc
    if rc != 0:
        print("Abortado: invariantes fallaron.")
        return rc
    out = "src/lib/services/plan-de-cuentas.data.ts"
    with open(out, "w", encoding="utf-8") as f:
        f.write(emit_ts(rows))
    print(f"\n✓ Escrito {out} ({len(rows)} cuentas)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
