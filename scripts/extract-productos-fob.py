#!/usr/bin/env python3
"""Extrae los productos del escopo (PCR/TBR/SUV/LTR/UHP) de la lista FOB de fábrica
y genera prisma/data/productos-fob.json para el importador TS.

Lee directamente el .xlsx (openpyxl). Mapea columnas POR NOMBRE de encabezado
(no por índice fijo) porque las hojas tienen layouts ligeramente distintos.
Solo lectura: no toca la base de datos.
"""
import json
import os
import re
import sys

import openpyxl

XLSX = os.path.expanduser(
    "~/Desktop/TABELA FOB/LISTA DE PRECIOS FOB - JULIO - ARGENTINA.xlsx"
)
OUT = os.path.join(os.path.dirname(__file__), "..", "prisma", "data", "productos-fob.json")

SCOPE = {"PCR", "TBR", "SUV", "SUV-NEW", "LTR", "LTR-NEW", "UHP"}

# Medida del neumático al inicio de la descripción. Cubre:
#   175/70R13 · 295/80R22.5 · 11R22.5 · 35X12.50R20LT · LT315/70R17 · 185R14C
#   195/70R15C · 205/50ZR16 (UHP, rating Z antes de la R)
MEDIDA_RE = re.compile(r"^(LT)?\d+(\.\d+)?([/X]\d+(\.\d+)?)?Z?R\d+(\.\d+)?[A-Z]{0,2}", re.I)

# Tokens de índice de carga/velocidad/ply a descartar al deducir el "modelo".
INDICE_RE = re.compile(r"^(\d+([/.]\d+)*[A-Z]{0,3}|\d+PR|XL|WL|TT|TTF|T/A|R/T|M\+S|M/S)$", re.I)

NULL_TOKENS = {"", "PENDING", "PENDIENTE", "PENDENTE", "·", "N/A", "-", "0"}


def clean(v):
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()


def collapse(s):
    return re.sub(r"\s+", " ", s).strip()


def to_num(v):
    s = clean(v)
    if s in NULL_TOKENS:
        return None
    try:
        n = float(s)
        return n if n > 0 else None
    except ValueError:
        return None


def to_int(v):
    n = to_num(v)
    return int(round(n)) if n is not None else None


def deducir_medida(desc):
    m = MEDIDA_RE.match(desc)
    return m.group(0).upper() if m else None


def deducir_modelo(desc, medida):
    """Mejor esfuerzo: quita la medida del inicio y los tokens de índice;
    lo que queda (palabras comerciales) es el modelo. None si queda vacío."""
    rest = desc
    if medida and rest.upper().startswith(medida.upper()):
        rest = rest[len(medida):]
    toks = rest.split()
    out = [t for t in toks if not INDICE_RE.match(t)]
    modelo = collapse(" ".join(out))
    return modelo or None


def main():
    if not os.path.exists(XLSX):
        print(f"ERROR: no se encontró el archivo:\n  {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    productos = []
    vistos = set()
    skipped_no_id = 0
    skipped_dup = 0
    por_categoria = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        col = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [clean(c) for c in row]
            # Localiza la fila de encabezado (la que contiene "ID" y "Description").
            if header is None:
                if "ID" in cells and "Description" in cells:
                    header = cells
                    col = {name: idx for idx, name in enumerate(cells) if name}
                continue
            # Salta encabezados repetidos en medio de la hoja.
            if cells and cells[0] == "ID":
                continue

            def get(name):
                idx = col.get(name)
                return cells[idx] if idx is not None and idx < len(cells) else ""

            categoria = get("Category")
            desc = collapse(get("Description"))
            if not desc or categoria not in SCOPE:
                continue

            codigo = get("ID").upper()
            if not codigo:
                skipped_no_id += 1
                continue
            if codigo in vistos:
                skipped_dup += 1
                continue
            vistos.add(codigo)

            medida = deducir_medida(desc)
            productos.append(
                {
                    "codigo": codigo,
                    "nombre": desc,
                    "marca": collapse(get("Brand")) or None,
                    "categoria": categoria,
                    "medida": medida,
                    "modelo": deducir_modelo(desc, medida),
                    "pesoNetoKg": to_num(get("Net Weight")),
                    "unidadesContenedor40hc": to_int(get("40 HC")),
                    "hoja": sheet,
                }
            )
            por_categoria[categoria] = por_categoria.get(categoria, 0) + 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    print(f"Extraídos: {len(productos)} productos -> {os.path.relpath(OUT)}")
    print(f"Por categoría: {por_categoria}")
    print(f"Saltados sin ID: {skipped_no_id} | duplicados de código: {skipped_dup}")
    sin_medida = [p["codigo"] for p in productos if not p["medida"]]
    print(f"Sin medida deducida: {len(sin_medida)}", sin_medida[:10])
    print("\nMuestra (8):")
    for p in productos[:8]:
        print(f"  {p['codigo']:>7} | {p['categoria']:<7} | medida={p['medida']!s:<14} | "
              f"modelo={p['modelo']!s:<22} | peso={p['pesoNetoKg']} | 40hc={p['unidadesContenedor40hc']} | {p['nombre'][:30]!r}")


if __name__ == "__main__":
    main()
